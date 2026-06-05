"""Excel exporter with multiple sheets, formatting, and conditional score colours."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from models.lead import Lead
from config import OUTPUT_DIR
from exporters.csv_exporter import HEADERS_PT, _format_value

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_GREEN_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")


class ExcelExporter:
    """Exports leads to a formatted .xlsx with three worksheets."""

    def save(
        self,
        leads: list[Lead],
        discarded: Optional[list[Lead]] = None,
        filename: str = "leads",
    ) -> str:
        """
        Save leads to output/{filename}_{timestamp}.xlsx.

        Sheets:
          1. Todos os Leads   – full dataset
          2. Leads Quentes    – score >= 60, sorted by score desc
          3. Descartados      – filtered-out leads with reason column
        """
        Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = Path(OUTPUT_DIR) / f"{filename}_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # delete the blank default sheet

        self._write_sheet(wb.create_sheet("Todos os Leads"), leads)

        hot = sorted(
            [l for l in leads if l.score_qualidade >= 60],
            key=lambda x: x.score_qualidade,
            reverse=True,
        )
        self._write_sheet(wb.create_sheet("Leads Quentes"), hot)

        self._write_sheet(
            wb.create_sheet("Descartados"),
            discarded or [],
            include_reason=True,
        )

        wb.save(filepath)
        logger.info("Excel saved: %s (%d leads)", filepath, len(leads))
        return str(filepath)

    def _write_sheet(
        self,
        ws,
        leads: list[Lead],
        include_reason: bool = False,
    ) -> None:
        """Write leads to a worksheet with headers, formatting, and conditional rules."""
        fields = list(HEADERS_PT.keys())
        headers = list(HEADERS_PT.values())

        if include_reason:
            fields = fields + ["filtro_motivo"]
            headers = headers + ["Motivo Descarte"]

        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 28

        # Write data rows
        for lead in leads:
            row = [_format_value(getattr(lead, field, "")) for field in fields]
            ws.append(row)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter on all columns
        ws.auto_filter.ref = ws.dimensions

        # Auto-fit column widths (capped at 60 chars)
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Conditional formatting on Score column
        if "Score" in headers and len(leads) > 0:
            score_idx = headers.index("Score") + 1
            score_letter = get_column_letter(score_idx)
            data_range = f"{score_letter}2:{score_letter}{len(leads) + 1}"

            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=_GREEN_FILL),
            )
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator="between", formula=["50", "79"], fill=_YELLOW_FILL),
            )
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator="lessThan", formula=["50"], fill=_RED_FILL),
            )
